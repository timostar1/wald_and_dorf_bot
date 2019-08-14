from dish_parser import dish_parser
from ingredients_calculator import ingredients_calculator
from openpyxl import load_workbook
from table_handler import first_blank_row_finder

def find_unit(ingredient):
    units = load_workbook(filename='units.xlsx')['Units']
    units_first_blank_row = first_blank_row_finder(units, 1)
    units_keys = [units.cell(row=i, column=1) for i in range(1, units_first_blank_row)]
    for cell in units_keys:
            if cell.value == ingredient:
                a = cell.row
                return units.cell(row=a,column=2).value
    return 'г'

def to_buy_handler(to_buy):
    answer = 'Чтобы приготовить эти блюда вам понадобится купить:'
    for ingredient in to_buy:
        answer += '\n' + ingredient + ' ' + str(to_buy[ingredient]) + find_unit(ingredient)
    return answer

def unknown_dishes_handler(unknown_dishes):
    if unknown_dishes:
        return 'Следующие блюда не были найдены в таблице блюд, поэтому мы не посчитали их: ' + ', '.join(unknown_dishes)
    else:
        return ''

def repeated_dishes_handler(repeated_dishes):
    if repeated_dishes:
        return 'Следующие блюда повторялись в вашем сообщении, поэтому мы посчитали их больше одного раза: ' + ', '.join(repeated_dishes)
    else:
        return ''

def message_composer(message_from_chief):
    to_cook, repeated_dishes = dish_parser(message_from_chief)
    if to_cook:
        to_buy, unknown_dishes = ingredients_calculator(to_cook)
        if to_buy:
            return to_buy_handler(to_buy) + '\n\n' + unknown_dishes_handler(unknown_dishes) + '\n' + repeated_dishes_handler(repeated_dishes)
        else:
            return 'К сожалению, мы не смогли найти блюда, которые вы написали, в таблице блюд'
    else:
        return 'К сожалению, мы не смогли найти блюд в вашем сообщении'