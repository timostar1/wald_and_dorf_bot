from openpyxl import load_workbook



def write_excel(category,item,place,day,username,phone_number,photo):
    '''
    записывает в excel новые экземпляры данных о вещах
    :param category: column=1
    :param day: column=4
    :param username: column=5
    :param item: column=2
    :param phone_number: column=6
    :param place: column=3
    :param photo: column=7
    :return:
    '''
    lost_and_found_file = load_workbook('lost and found.xlsx')
    lost_and_found_page = lost_and_found_file["DB"]

    row = 2
    a = lost_and_found_page.cell(row=2, column=1).value

    while a != None:
        row += 1
        a = lost_and_found_page.cell(row=row, column=1).value

    m = [category, item, place, username, day, phone_number,photo]
    for i in range (1,8):

        lost_and_found_page.cell(row=row, column=i).value = m[i-1]



    lost_and_found_file.save('lost and found.xlsx')
    # print(row)

write_excel('категория', 'название', 'место потери', 'имя пользователя', 'дата', 'номер телефона', 'фото')
