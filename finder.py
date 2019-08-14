from openpyxl import load_workbook


def search_item(categ):
    '''
    команда по которой ищется определенная вещь
    :param categ: категория которую задает пользователь
    :return:
    '''
    wb2 = load_workbook('lost and found.xlsx')
    # wb2 - переменная отвечающая за Excel (workbook)
    ws = wb2["DB_test"]
    # ws - переменная,отвечающая за рабочий лист Excel ( DB_test )
    i = 2
    # i - счетчик рядов
    cell = ws.cell(row=i, column=1).value
    while cell is not None:
        if categ == cell:
            m = []
            # m - список параметров вещи
            for j in range(1, 7):
                # j - счетчик колонок
                m.append(ws.cell(row=i, column=j).value)
            print(i, *m)
        i += 1
        cell = ws.cell(row=i, column=1).value


search_item("Одежда")
