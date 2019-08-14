from openpyxl import load_workbook

def read_numbers():
    with open('numbers.txt', 'r') as file:
        string = file.read()
        number_of_people, number_of_vegans = [int(i) for i in string.split("\n")]
        return number_of_people, number_of_vegans
number_of_people, number_of_vegans = read_numbers()

def ingredients_calculator(to_cook):
    '''
    :param to_cook:
    :return: to_buy, unknown_dishes
    '''

    # Функция получает на вход список блюд (словарь to_cook)
    # Также она обращается к файлу numbers.txt,
    # В котором хранится количчество людей и количество вегеарианцев
    # Ещё она обращается к excel-таблице ingredients,
    # В которой хранятся рецепты блюд
    # На выходе она выдаст список продуктов для покупки (словарь to_buy)
    # И, возможно, список неизвестных ей блюд (unknown_dishes),
    # Т. е. тех, которые не были найдены в таблице ingredients

    to_buy = dict()
    unknown_dishes = set()

    # Импорт таблицы с граммовками
    ingredients = load_workbook(filename='ingredients.xlsx')['Ingredients']

    # Цикл ниже находит первую пустую строку в таблице
    # и записывает её номер как first_blank_row
    value = True
    first_blank_row = 1
    while value != None:
        value = ingredients.cell(row=first_blank_row, column=3).value
        first_blank_row += 1
    first_blank_row -= 1

    # В dish_names запишем все названия блюд
    dish_names = [ingredients.cell(row=i, column=1) for i in range(1, first_blank_row)]

    # Функция ниже находит нужное название блюда в списке dish_names и выдаёт его строку
    def search_dish(name):
        for cell in dish_names:
            if cell.value == name:
                return cell.row
        return 'not_found'

    # Эта функция получает на вход строку блюда.
    # Она показывает, сколько строк ингредиентов относится к этому блюду
    def search_next_dish(row):
        end = 1
        value = None
        while (value == None) and (row + end <= first_blank_row):
            value = ingredients.cell(column=1, row=row+end).value
            end += 1
        return end

    # Для каждого блюда в списке to_cook
    for dish in to_cook:
        row = search_dish(dish)
        # Находим его ряд
        if row == 'not_found':
            # Если блюда нет в таблице ingredients
            # То мы добавим его в список ненайденных блюд
            unknown_dishes.add(dish)
        else:
            # Если же оно там есть, то
            # умножим каждый ингредиент на число людей и количество блюда
            for i in range(row, row + search_next_dish(row) - 1):
                ingredient = ingredients.cell(column=2,row=i).value
                quantity = ingredients.cell(column=3,row=i).value * to_cook[dish] * number_of_people
                if ingredient in to_buy:
                    # Если этот ингредиент уже есть в словаре
                    # Мы сложим количество этого ингредиента с уже имеющимся
                    to_buy[ingredient] += quantity
                else:
                    # Если нет, то добавим этот ингредиент в словарь
                    to_buy[ingredient] = quantity

    return to_buy, unknown_dishes,